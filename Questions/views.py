from django.shortcuts import render, HttpResponseRedirect
from django.http import JsonResponse
from Questions.models import *
from Category.models import Category
from openpyxl import load_workbook
from django.views.generic import View
from .forms import ExcelForm
from PalestineHistory.settings import *
from django.core.urlresolvers import reverse
from Client.models import UserAgents


class ViewQuestions(View):
    def get(self, request):
        if request.session.get('manage', None):
            session_user = request.session.get('manage', None)
            user_agent_oop = UserAgents.objects.filter(user=session_user).latest('id')
            if not user_agent_oop.logout:
                all_questions = Question.objects.all()
                context = {
                    'title': 'الأسئلة',
                    'Questions': all_questions,
                    'logout_user': session_user,
                    'admin_is_manager': request.session.get('is_manager', False),
                    'session_user': session_user.split('_')[0],
                }
                return render(self.request, 'Dashboard/Questions/view_questions.html', context)
        return HttpResponseRedirect(reverse('Dashboard:login'))


class UploadFiles(View):
    def get(self, request):
        if request.session.get('manage', None):
            session_user = request.session.get('manage', None)
            user_agent_oop = UserAgents.objects.filter(user=session_user).latest('id')
            if not user_agent_oop.logout:
                excel_list = ExcelFiles.objects.all()
                context = {
                    'title': 'رفع الأسئلة',
                    'excel_list': excel_list,
                    'logout_user': session_user,
                    'admin_is_manager': request.session.get('is_manager', False),
                    'session_user': session_user.split('_')[0],
                }
                return render(self.request, 'Dashboard/Questions/upload_excel.html', context)
        return HttpResponseRedirect(reverse('Dashboard:login'))


class UploadExcel(View):
    def post(self, request):
        if request.session.get('manage', None):
            session_user = request.session.get('manage', None)
            user_agent_oop = UserAgents.objects.filter(user=session_user).latest('id')
            if not user_agent_oop.logout:
                form = ExcelForm(self.request.POST, self.request.FILES)
                if form.is_valid():
                    excel = form.save()
                    url = excel.file.url
                    name = excel.file.name
                    context = read_excel(url, name.split('/')[1])
                    if str(context) != 'empty':
                        data = {'is_valid': True, 'name': name, 'url': url, 'notes': context}
                        return JsonResponse(data)
                return JsonResponse({'is_valid': False})
        return HttpResponseRedirect(reverse('Dashboard:login'))


def read_excel(url, name):
    workbook = load_workbook(filename=BASE_DIR + str(url), read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    if sheet.rows:
        category_oop = Category.objects.filter(pk=1).first()
        wrong_question = dict()
        row_number = valid_question = 0
        for row in sheet.rows:
            row_number += 1
            list_q = []
            for cell in range(5):
                if row[cell].value is not None:
                    list_q.append(row[cell].value)
                else:
                    if name in wrong_question:
                        if row_number not in wrong_question[name]:
                            wrong_question[name].append(row_number)
                    else:
                        wrong_question[name] = [row_number]
            if len(list_q) == 5:
                question_oop = Question.objects.filter(category=category_oop, question=list_q[0],
                                                       answer_1=list_q[1], answer_2=list_q[2],
                                                       answer_3=list_q[3], answer_4=list_q[4]).first()
                if not question_oop:
                    valid_question += 1
                    Question(category=category_oop, question=list_q[0], answer_1=list_q[1], answer_2=list_q[2],
                             answer_3=list_q[3], answer_4=list_q[4]).save()
        context = {
            'all_question': len(Question.objects.filter(category=category_oop)),
            'add__question': valid_question,
            'wrong_question': wrong_question
        }
        workbook.close()
        os.remove(BASE_DIR + str(url))
        return context
    else:
        return 'empty'
